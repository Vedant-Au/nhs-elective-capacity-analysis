"""
Loader for NHS Workforce Statistics (HCHS staff), one March snapshot per
fiscal year — same annual-snapshot scope decision as GPAD, and for the same
reason: this is a contextual/secondary signal for the "why is the system
under pressure" narrative (staffing capacity alongside bed/diagnostic
capacity), not an input the elective pressure index strictly needs at
monthly resolution.

Restricted to the 13 Cheshire and Merseyside NHS trust codes in
sql/reference/cheshire_merseyside_providers.csv (independent-sector and GP
providers in that file aren't HCHS staff and wouldn't appear in this
source anyway) rather than the wider NHS England region/ICS breakdown, to
keep this at the same provider grain as RTT/DM01/A&E/KH03.

Column layout drifted across the window in ways that matter and had to be
handled explicitly, not assumed stable. Four genuinely different variants
turned up across 7 files, not the two I expected going in:
  - March 2020 / March 2021: no Integrated Care System columns (ICSs
    didn't exist as a formal NHS tier until July 2022), and no header text
    at all on the org name/code columns — confirmed by inspection, not
    assumed. Blank spacer columns sit between each staff-group subtotal.
  - March 2022: still no ICS columns and still no header text on org
    name/code, BUT the blank spacer columns between staff-group subtotals
    are gone — a hybrid of the two adjacent years' layouts, discovered
    only because the naive two-era assumption made on the first pass
    through this loader failed with a "could not locate header row" error
    on this exact file and forced a closer look.
  - March 2023 - March 2025: ICS code/name columns inserted before
    Organisation name/code, full header text on every column, no spacers.
  - March 2026: sheet renamed from '3. NHSE, Org & SG - FTE' to '2', file
    itself renamed from "... England and Organisation" to "... Trusts and
    core organisations - data tables", Data month/Data type columns added
    before NHS England region code — otherwise same semantics as 2023-25.

Rather than hardcode four separate index maps (which is exactly the kind
of thing that breaks silently if a future release shuffles columns again),
every file is read with the SAME anchor-and-search resolver: 'Nurses &
health visitors' and 'Scientific, therapeutic & technical staff' are the
one pair of header labels confirmed identical, unambiguous, and present in
all four variants (including the two "unlabelled" years — only the org
name/code columns lack header text there, not the staff-group columns).
Every other needed column is located by searching backward from that
anchor for the nearest matching label, which naturally resolves Era C's
repeated 'All staff groups' text to the correct one of its three
occurrences (nearest-preceding-match, not first-match). Organisation code
and name are then read as fixed offsets (-1, -2) from wherever 'Total' /
'All staff groups' resolves, which held across all four variants once
checked directly against real data rows — not assumed to hold by pattern.

Scope decision, flagged same as elsewhere in this warehouse: only Total
FTE, Professionally Qualified Clinical Staff, HCHS Doctors (all grades),
Consultants, Nurses & Health Visitors, and Scientific/Therapeutic &
Technical staff (the group that includes radiographers — directly
relevant to DM01 diagnostic capacity) are staged. The finer sub-grade
breakdown (Associate Specialist/Specialty Doctor/registrars/foundation
years, support-staff breakdown, infrastructure breakdown) is available in
the same source files if a later phase needs it — deferred, not discarded.

Usage:
    import duckdb
    from load_workforce import load_workforce
    con = duckdb.connect('nhs_warehouse.db')
    con.execute(open('sql/schema/12_staging_workforce.sql').read())
    load_workforce(con, workforce_dir='data_raw/workforce')
    con.execute(open('sql/schema/13_fact_workforce_provider_year.sql').read())
"""

import openpyxl
from datetime import date

FILES = {
    'workforce_2019-20_march2020.xlsx': ('2019-20', date(2020, 3, 31)),
    'workforce_2020-21_march2021.xlsx': ('2020-21', date(2021, 3, 31)),
    'workforce_2021-22_march2022.xlsx': ('2021-22', date(2022, 3, 31)),
    'workforce_2022-23_march2023.xlsx': ('2022-23', date(2023, 3, 31)),
    'workforce_2023-24_march2024.xlsx': ('2023-24', date(2024, 3, 31)),
    'workforce_2024-25_march2025.xlsx': ('2024-25', date(2025, 3, 31)),
    'workforce_2025-26_march2026.xlsx': ('2025-26', date(2026, 3, 31)),
}

# Sheet holding the org-level FTE table changed name once, in the March
# 2026 release (see module docstring) — everything else is on
# '3. NHSE, Org & SG - FTE'.
_CANDIDATE_SHEETS = ['3. NHSE, Org & SG - FTE', '2']

# Cheshire and Merseyside NHS trust codes only (excludes the independent-
# sector and GP-level rows in the reference CSV, which aren't HCHS staff
# and wouldn't appear in this source). RW4 (Mersey Care, mental health) is
# included here even though it's flagged include_in_core_analysis=0 in the
# reference file for the acute-elective fact tables — workforce is a
# broader contextual signal than the elective pressure index itself.
TARGET_CODES = {'RBL', 'RBN', 'RBQ', 'RBS', 'RBT', 'REM', 'REN', 'REP', 'RET', 'RJN', 'RJR', 'RW4', 'RWW'}

def _find_sheet(wb):
    for name in _CANDIDATE_SHEETS:
        if name in wb.sheetnames:
            return wb[name]
    raise ValueError(f"None of {_CANDIDATE_SHEETS} found in workbook (sheets present: {wb.sheetnames})")


def _resolve_columns(header_row):
    """Anchor on 'Nurses & health visitors' — the one label confirmed
    identical and unambiguous across all four column-layout variants seen
    in this source — then search backward for everything else. Backward
    search resolves repeated labels (Era C's three separate 'All staff
    groups' subtotals) to the nearest preceding match rather than the
    first occurrence in the row, which is what actually lines up with the
    real data in every variant checked."""
    def find_backward(predicate, start):
        for i in range(start, -1, -1):
            v = header_row[i]
            if v and predicate(str(v).strip()):
                return i
        raise ValueError(f"Could not find expected column searching backward from {start} in header: {header_row}")

    nurses_idx = None
    for i, v in enumerate(header_row):
        if v and str(v).strip() == 'Nurses & health visitors':
            nurses_idx = i
            break
    if nurses_idx is None:
        raise ValueError(f"Could not find 'Nurses & health visitors' anchor column in header: {header_row}")

    scitech_idx = None
    for i in range(nurses_idx, len(header_row)):
        v = header_row[i]
        if v and str(v).strip() == 'Scientific, therapeutic & technical staff':
            scitech_idx = i
            break
    if scitech_idx is None:
        raise ValueError(f"Could not find 'Scientific, therapeutic & technical staff' column in header: {header_row}")

    consultant_idx = find_backward(lambda v: 'Consultant' in v and 'Associate' not in v, nurses_idx - 1)
    hchs_doctors_idx = find_backward(lambda v: 'HCHS Doctors' in v or 'HCHS doctors' in v, consultant_idx - 1)
    prof_qual_idx = find_backward(lambda v: v in ('Professionally qualified clinical staff', 'All staff groups'), hchs_doctors_idx - 1)
    total_idx = find_backward(lambda v: v in ('Total', 'All staff groups'), prof_qual_idx - 1)

    return {
        'org_code': total_idx - 1, 'org_name': total_idx - 2, 'total_fte': total_idx,
        'prof_qual_clinical_fte': prof_qual_idx, 'hchs_doctors_fte': hchs_doctors_idx,
        'consultant_fte': consultant_idx, 'nurses_fte': nurses_idx, 'scitech_fte': scitech_idx,
    }


def _load_one_file(con, fname, fiscal_year, period_end_date, workforce_dir):
    wb = openpyxl.load_workbook(f'{workforce_dir}/{fname}', read_only=True, data_only=True)
    ws = _find_sheet(wb)

    header_row_idx, header_row = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if any(v and str(v).strip() == 'Nurses & health visitors' for v in row):
            header_row_idx, header_row = i, row
            break
    if header_row is None:
        raise ValueError(f"{fname}: could not locate a header row containing 'Nurses & health visitors'")
    cols = _resolve_columns(header_row)
    data_start_row = header_row_idx + 1

    n = 0
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if cols['org_code'] >= len(row):
            continue
        org_code = row[cols['org_code']]
        if org_code not in TARGET_CODES:
            continue
        con.execute(
            """insert into stg_workforce_raw
               (fiscal_year, period_end_date, provider_org_code, provider_org_name,
                total_fte, prof_qual_clinical_fte, hchs_doctors_fte, consultant_fte,
                nurses_health_visitors_fte, scientific_therapeutic_technical_fte,
                source_file)
               values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [fiscal_year, period_end_date, org_code, row[cols['org_name']],
             row[cols['total_fte']], row[cols['prof_qual_clinical_fte']],
             row[cols['hchs_doctors_fte']], row[cols['consultant_fte']],
             row[cols['nurses_fte']], row[cols['scitech_fte']], fname],
        )
        n += 1
    return n


def load_workforce(con, workforce_dir='data_raw/workforce'):
    """Load all 7 workforce annual snapshots (FY2019-20 through FY2025-26,
    Cheshire and Merseyside trusts only) into stg_workforce_raw."""
    total = 0
    for fname, (fiscal_year, period_end_date) in FILES.items():
        n = _load_one_file(con, fname, fiscal_year, period_end_date, workforce_dir)
        total += n
        if n != len(TARGET_CODES):
            print(f"WARNING: {fname} matched {n}/{len(TARGET_CODES)} target trusts — investigate before trusting this year's data")
    print(f"Workforce: loaded {total} rows across {len(FILES)} fiscal years ({len(TARGET_CODES)} trusts targeted per year)")
    return total
