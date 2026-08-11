"""
Loader for IMD 2019 (English Indices of Deprivation), Cheshire and
Merseyside's 9 upper-tier local authorities only.

Unlike every other source in this warehouse, IMD 2019 is NOT a time
series — it's a single point-in-time publication (26 Sep 2019, methodology
based on data mostly from 2015-2017) with no equivalent annual refresh
during this project's FY2019-20-FY2025-26 window (the next full IMD update
wasn't published until 2025, outside this project's original data-pull).
So there's no fiscal_year dimension here — this loads once into a static
reference table, not a fact table with a time grain.

Source is gov.uk's File 11 (upper-tier local authority district
summaries) rather than the LSOA-level File 1 that the publication's own
guidance calls "usually sufficient for most users" — LSOA level (32,844
areas nationally) is far finer than anything else in this warehouse joins
against; nothing else here has an LSOA key. Upper-tier local authority is
the right grain: interpretable on its own, small (151 LAs nationally, 9 of
them Cheshire and Merseyside), and exactly what the original charter's
inequality-of-access rationale for choosing this ICB was pointing at
(Liverpool/Knowsley vs the more affluent Cheshire boroughs).

Scope decision, flagged same as elsewhere: loads the IMD sheet (overall
Index of Multiple Deprivation) and the Health sheet (Health Deprivation
and Disability domain — the one domain most directly relevant to a
healthcare capacity project) out of File 11's 11 sheets. The other 7
domains (Income, Employment, Education, Crime, Barriers to Housing and
Services, Living Environment, plus the IDACI/IDAOPI supplementary indices)
are sitting in the same downloaded file and are a legitimate Phase 2+
enhancement for a fuller socioeconomic narrative — not loaded now because
this project's use case (deprivation as a lens on the elective/diagnostic
pressure story) doesn't need all nine cuts to make the point.

NOT done here, deliberately: mapping providers/trusts to a single "home"
local authority. Trust catchments cross LA boundaries (e.g. Liverpool
University Hospitals draws patients from well beyond Liverpool itself),
so which LA's deprivation score to attach to which provider is an
analytical modelling choice, not a warehousing fact — that decision
belongs in the analytics layer, made deliberately, not baked in silently
here.

Usage:
    import duckdb
    from load_imd import load_imd
    con = duckdb.connect('nhs_warehouse.db')
    con.execute(open('sql/schema/14_dim_imd2019_local_authority.sql').read())
    load_imd(con, imd_dir='data_raw/imd')
"""

import openpyxl

FILE = 'imd2019_upper_tier_la_summaries.xlsx'

# The 9 upper-tier local authorities making up NHS Cheshire and Merseyside
# ICB's geography, by 2019 ONS code.
TARGET_LA_CODES = {
    'E06000006',  # Halton
    'E06000007',  # Warrington
    'E06000049',  # Cheshire East
    'E06000050',  # Cheshire West and Chester
    'E08000011',  # Knowsley
    'E08000012',  # Liverpool
    'E08000013',  # St. Helens
    'E08000014',  # Sefton
    'E08000015',  # Wirral
}


def load_imd(con, imd_dir='data_raw/imd'):
    wb = openpyxl.load_workbook(f'{imd_dir}/{FILE}', read_only=True, data_only=True)

    imd_by_la = {}
    ws = wb['IMD']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] in TARGET_LA_CODES:
            imd_by_la[row[0]] = {
                'la_name': row[1],
                'imd_avg_rank': row[2], 'imd_rank_of_avg_rank': row[3],
                'imd_avg_score': row[4], 'imd_rank_of_avg_score': row[5],
                'imd_pct_lsoas_most_deprived_decile': row[6],
                'imd_rank_of_pct_most_deprived_decile': row[7],
            }

    health_by_la = {}
    ws = wb['Health']
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] in TARGET_LA_CODES:
            health_by_la[row[0]] = {
                'health_avg_rank': row[2], 'health_rank_of_avg_rank': row[3],
                'health_avg_score': row[4], 'health_rank_of_avg_score': row[5],
                'health_pct_lsoas_most_deprived_decile': row[6],
                'health_rank_of_pct_most_deprived_decile': row[7],
            }

    n = 0
    for la_code in TARGET_LA_CODES:
        if la_code not in imd_by_la:
            raise ValueError(f"Expected LA code {la_code} not found in IMD sheet — check TARGET_LA_CODES against the source file")
        if la_code not in health_by_la:
            raise ValueError(f"Expected LA code {la_code} not found in Health sheet — check TARGET_LA_CODES against the source file")
        i, h = imd_by_la[la_code], health_by_la[la_code]
        con.execute(
            """insert into dim_imd2019_local_authority
               (la_code, la_name,
                imd_avg_rank, imd_rank_of_avg_rank, imd_avg_score, imd_rank_of_avg_score,
                imd_pct_lsoas_most_deprived_decile, imd_rank_of_pct_most_deprived_decile,
                health_avg_rank, health_rank_of_avg_rank, health_avg_score, health_rank_of_avg_score,
                health_pct_lsoas_most_deprived_decile, health_rank_of_pct_most_deprived_decile,
                source_file)
               values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [la_code, i['la_name'],
             i['imd_avg_rank'], i['imd_rank_of_avg_rank'], i['imd_avg_score'], i['imd_rank_of_avg_score'],
             i['imd_pct_lsoas_most_deprived_decile'], i['imd_rank_of_pct_most_deprived_decile'],
             h['health_avg_rank'], h['health_rank_of_avg_rank'], h['health_avg_score'], h['health_rank_of_avg_score'],
             h['health_pct_lsoas_most_deprived_decile'], h['health_rank_of_pct_most_deprived_decile'],
             FILE],
        )
        n += 1
    print(f"IMD 2019: loaded {n}/{len(TARGET_LA_CODES)} Cheshire and Merseyside local authorities")
    return n
