"""
Loader for GPAD (Appointments in General Practice), one March snapshot per
fiscal year — see sql/schema/10_staging_gpad.sql header for why this is an
annual snapshot rather than all 84 monthly files.

Reads Table 3a (appointments by status) and Table 4 (appointments + patient
list size) out of each year's "Summary" xlsx and inserts both into
stg_gpad_raw, restricted to National / Region / STP-or-ICB tier geography
rows (identified by ons_code prefix — 'ENG', 'E40xxxxxx', 'E54xxxxxx' —
since the "Type" label text itself changes across years and isn't a
reliable filter: 2020 has 'Regional Local Office' and 'STP' rows covering
the same geography, 2022-23 onwards relabels the STP tier as 'ICB').

Confirmed by inspection on 2026-08-07 that both tables have an identical
header row (13) / data-start row (14 for Table 3a, 12 for Table 4)
structure in all 7 files, despite sheet-count and filename drift elsewhere
in the workbook (e.g. 2023-24 onwards adds Table 1a/1b/1c and a PCN
Information sheet that this loader doesn't touch).

Usage:
    import duckdb
    from load_gpad import load_gpad
    con = duckdb.connect('nhs_warehouse.db')
    con.execute(open('sql/schema/10_staging_gpad.sql').read())
    load_gpad(con, gpad_dir='data_raw/gpad')
    con.execute(open('sql/schema/11_fact_gpad_geography_year.sql').read())
"""

import openpyxl
from datetime import date

# filename -> (fiscal_year, period_end_date). Fiscal year label follows the
# Apr-Mar convention used throughout this project: the March 2020 snapshot
# is the year-end point of FY2019-20, not FY2020-21.
FILES = {
    'gpad_2019-20_march2020.xlsx': ('2019-20', date(2020, 3, 31)),
    'gpad_2020-21_march2021.xlsx': ('2020-21', date(2021, 3, 31)),
    'gpad_2021-22_march2022.xlsx': ('2021-22', date(2022, 3, 31)),
    'gpad_2022-23_march2023.xlsx': ('2022-23', date(2023, 3, 31)),
    'gpad_2023-24_march2024.xlsx': ('2023-24', date(2024, 3, 31)),
    'gpad_2024-25_march2025.xlsx': ('2024-25', date(2025, 3, 31)),
    'gpad_2025-26_march2026.xlsx': ('2025-26', date(2026, 3, 31)),
}


def _keep(ons_code):
    """National / Region / STP-or-ICB tier only — excludes CCG, Sub ICB
    Location, and the pre-2022 'Regional Local Office' tier (E39xxxxxx),
    which sits alongside STP at the same level in the 2020 file and would
    double-count Cheshire and Merseyside's appointments if not excluded."""
    if ons_code is None:
        return False
    return ons_code == 'ENG' or ons_code.startswith('E40') or ons_code.startswith('E54')


def _load_table_3a(con, wb, fname, fiscal_year, period_end_date):
    ws = wb['Table 3a']
    n = 0
    for row in ws.iter_rows(min_row=14, values_only=True):
        if row[0] is None:
            break  # blank row marks end of data, before footnotes/copyright
        geog_type_raw, nhs_area_code, ons_code, name = row[0], row[1], row[2], row[3]
        if not _keep(ons_code):
            continue
        open_active, included, total, attended, dna, unknown = (
            row[4], row[5], row[6], row[8], row[9], row[10]
        )
        con.execute(
            """insert into stg_gpad_raw
               (fiscal_year, period_end_date, source_table, geog_type_raw,
                nhs_area_code, ons_code, geog_name, open_active_practices,
                included_practices, appointments_total, appointments_attended,
                appointments_dna, appointments_unknown, source_file)
               values (?, ?, 'Table 3a', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            [fiscal_year, period_end_date, geog_type_raw, nhs_area_code, ons_code,
             name, open_active, included, total, attended, dna, unknown, fname],
        )
        n += 1
    return n


def _load_table_4(con, wb, fname, fiscal_year, period_end_date):
    ws = wb['Table 4']
    n = 0
    for row in ws.iter_rows(min_row=12, values_only=True):
        if row[0] is None:
            break
        geog_type_raw, nhs_area_code, ons_code, name = row[0], row[1], row[2], row[3]
        if not _keep(ons_code):
            continue
        total, list_size = row[4], row[6]
        con.execute(
            """insert into stg_gpad_raw
               (fiscal_year, period_end_date, source_table, geog_type_raw,
                nhs_area_code, ons_code, geog_name, appointments_total,
                patient_list_size, source_file)
               values (?, ?, 'Table 4', ?, ?, ?, ?, ?, ?, ?)""",
            [fiscal_year, period_end_date, geog_type_raw, nhs_area_code, ons_code,
             name, total, list_size, fname],
        )
        n += 1
    return n


def load_gpad(con, gpad_dir='data_raw/gpad'):
    """Load all 7 GPAD annual snapshots (FY2019-20 through FY2025-26) into
    stg_gpad_raw on the given DB connection (DuckDB or psycopg2-style)."""
    total_3a, total_4 = 0, 0
    for fname, (fiscal_year, period_end_date) in FILES.items():
        wb = openpyxl.load_workbook(f'{gpad_dir}/{fname}', read_only=True, data_only=True)
        total_3a += _load_table_3a(con, wb, fname, fiscal_year, period_end_date)
        total_4 += _load_table_4(con, wb, fname, fiscal_year, period_end_date)
    print(f"GPAD: loaded {total_3a} rows from Table 3a, {total_4} rows from Table 4 across {len(FILES)} fiscal years")
    return total_3a + total_4
