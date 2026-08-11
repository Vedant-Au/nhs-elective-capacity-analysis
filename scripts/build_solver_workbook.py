"""
Step 3: build the actual Excel Solver workbook from /tmp/solver_model_solution.json
and /tmp/budget_sensitivity.json. All computed cells are live formulas (SUM,
SUMIF, MIN), not hardcoded Python results, per this project's xlsx standard.
Decision variable cells are pre-filled with the Python-validated reference-
optimal solution so the workbook opens already showing an optimized plan,
but every decision cell is editable and Solver-ready for Vedant to re-solve
under different assumptions.

Row references into the Assumptions sheet are resolved through the `AROW`
dict by name, not hardcoded — a hand-counted row number was exactly the
kind of silent-error risk this project's own conventions warn against
(same category of mistake as the imd_avg_rank/imd_rank_of_avg_rank mixup
in the inequality-of-access phase), so building it this way instead of
patching individual `$C$NN` guesses after the fact.

No VBA macro is included. openpyxl cannot author a new VBA project (only
preserve one already in a template file, via keep_vba=True), and there is
no existing macro-enabled template to start from in this project — shipping
a hand-crafted vbaProject.bin blind, with no way to test it actually runs
Excel's Solver correctly, would risk a broken file. A plain .xlsx with exact
Solver dialog instructions (cell-by-cell) is the honest choice here over an
untested macro. Flagged, not silently worked around.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

with open('/tmp/solver_model_solution.json') as f:
    bundle = json.load(f)
with open('/tmp/budget_sensitivity.json') as f:
    sens = json.load(f)

providers = bundle['providers']
providers.sort(key=lambda p: p['provider_org_code'])
n_months = bundle['meta']['n_months']
a = bundle['assumptions']
ref = {r['provider_org_code']: r for r in bundle['reference_solution']}

ARIAL = 'Arial'
F_TITLE = Font(name=ARIAL, size=14, bold=True, color='1F4E78')
F_HEADER = Font(name=ARIAL, size=10, bold=True, color='FFFFFF')
F_LABEL = Font(name=ARIAL, size=10, bold=True)
F_NORMAL = Font(name=ARIAL, size=10)
F_INPUT = Font(name=ARIAL, size=10, color='0000FF')
F_FORMULA = Font(name=ARIAL, size=10, color='000000')
F_LINK = Font(name=ARIAL, size=10, color='008000')
F_NOTE = Font(name=ARIAL, size=9, italic=True, color='808080')
FILL_HEADER = PatternFill('solid', fgColor='1F4E78')
FILL_ASSUMPTION = PatternFill('solid', fgColor='FFFF00')
FILL_TOTAL = PatternFill('solid', fgColor='D9E1F2')
FILL_WARN = PatternFill('solid', fgColor='FCE4D6')
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CUR = '£#,##0;(£#,##0);"-"'
PCT = '0.0%'

wb = openpyxl.Workbook()

# ===========================================================================
# SHEET 1: Read Me
# ===========================================================================
ws = wb.active
ws.title = 'Read Me'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 110

r = 2
ws.cell(r, 2, 'NHS Cheshire & Merseyside — Elective Capacity Optimizer').font = F_TITLE
r += 2
intro = [
 ("Objective", "Minimize the total forecasted number of RTT incomplete pathways waiting "
  "over 52 weeks, summed across the 12 core Cheshire & Merseyside trusts, over the "
  f"{n_months}-month forecast horizon ({bundle['meta']['horizon_start']} to {bundle['meta']['horizon_end']})."),
 ("Decision variables", "36 cells (12 providers x 3 levers): additional monthly in-house "
  "elective capacity, independent-sector outsourcing, and diagnostic capacity investment, "
  "each expressed in RTT-equivalent completed pathways/month so the three levers are directly "
  "comparable. See the 'Solver Model' tab."),
 ("Grain decision", "Decision variables sit at PROVIDER level, not provider-specialty. Two of "
  "the three capacity-ceiling data sources (NHS Workforce Statistics, KH03 bed occupancy) only "
  "exist at provider grain in the underlying warehouse — there is no specialty-level FTE or bed "
  "data to constrain against — so a genuinely specialty-disaggregated model isn't supportable "
  "by the data actually available, independent of Excel's Solver also having a practical "
  "~200-variable ceiling that a 12 x ~20-specialty x 3-lever model would exceed anyway."),
 ("Constraints", "(1) Total cost <= budget envelope (an adjustable scenario input, not an "
  "asserted real ICB budget — see Assumptions tab); (2) each lever capped per provider by "
  "workforce/bed headroom or current activity growth rate; (3) independent-sector cap bounded "
  "by the ICB's current real outsourcing volume; (4) a BOUNDED equity tolerance — higher-"
  "deprivation providers must receive at least their baseline breach share minus 15 "
  "percentage points, not a hard no-worsening floor. All four per Vedant's confirmed design, "
  "docs/STATUS.md, 2026-08-07."),
 ("How to use this workbook", "The 'Solver Model' tab is pre-filled with a Python-validated "
  "reference-optimal solution (scipy.optimize.linprog, HiGHS solver) for the default £15m "
  "budget scenario. Every decision cell (blue, columns I:K) is live and Solver-ready — change "
  "the budget or any Assumption and re-run Excel's own Solver (Data > Solver) to re-optimize; "
  "exact cell references are on the 'Solver Model' tab itself."),
 ("What's genuinely sourced vs. estimated", "UPDATED 2026-08-08, twice. First: Vedant supplied "
  "screenshots of NHS England's National Cost Collection (NCC) 2024/25 Power BI dashboard "
  "(this sandbox still can't pull it directly — web_fetch returns the underlying file as "
  "unusable opaque binary). The blended tariff is built from REAL national unit costs "
  "(Elective Inpatient £6,624, Daycase £1,078, Outpatient Procedures £233) combined with this "
  "ICB's own REAL admitted/non-admitted completion mix (14.7%/85.3%, from the warehouse "
  "itself), and the diagnostic lever's base cost from a real DM01-test-mix-weighted blend of "
  "four NCC diagnostic currencies (imaging, audiology, other physiological measurement, "
  "endoscopy-as-outpatient-procedure). Second: the ERF 75%/100% marginal-rate split — flagged "
  "as unverified ever since it was first used — was checked against the actual 2025/26 NHS "
  "Payment Scheme and found WRONG. That mechanism was removed for 2025/26: NHS and "
  "independent-sector providers are now both paid 100% of NHSPS unit price for elective "
  "activity, no marginal rates. c1 (in-house) and c2 (outsourcing) are corrected to be equal. "
  "Remaining unsourced figures: the diagnostic tests-per-pathway conversion (1.75x) and the "
  "one bridging assumption where RTT's own data can't distinguish daycase from inpatient "
  "completions. Full derivation, every source cell, is on the Assumptions tab."),
 ("A finding worth reading before trusting one budget number", "With in-house and outsourcing "
  "now correctly priced EQUALLY (both £472/unit, following the 2025/26 payment-scheme "
  "correction above), the earlier 'outsourcing only activates once in-house is also full' "
  "story simplifies further: at any budget where both levers have spare capacity, the model is "
  "genuinely indifferent between them — Solver will fill whichever has capacity headroom left, "
  "not prefer one on cost grounds. Diagnostics (£247/unit) still fills first since it's "
  "structurally cheaper. The 'Budget Sensitivity' tab shows in-house/outsourcing activating "
  "together once diagnostic headroom exhausts (~£90m ICB-wide), with total reduction "
  "plateauing near 98% past that point — a capacity ceiling, not a budget ceiling."),
]
for label, body in intro:
    ws.cell(r, 2, label).font = F_LABEL
    r += 1
    c = ws.cell(r, 2, body)
    c.font = F_NORMAL
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 15 * (len(body) // 100 + 2)
    r += 2

# ===========================================================================
# SHEET 2: Assumptions  (build row-by-row, tracking real row numbers in AROW)
# ===========================================================================
ws = wb.create_sheet('Assumptions')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 70

ws.cell(2, 2, 'Model Assumptions').font = F_TITLE
ws.cell(4, 2, 'Yellow = key input, edit here (formulas elsewhere reference these cells)').font = F_NOTE

headers = ['Assumption', 'Value', 'Source / flag']
for i, h in enumerate(headers):
    c = ws.cell(6, 2 + i, h); c.font = F_HEADER; c.fill = FILL_HEADER

# (name, value_or_None, number_format, is_hardcoded_input, note)
# value=None means it's a formula, filled in a second pass below once all rows are known.
spec = [
    ('ncc_elective_inpatient_cost', a['ncc_elective_inpatient_cost'], CUR, True,
     'SOURCED — NHS England National Cost Collection (NCC) 2024/25, National Schedule of NHS '
     'costs, Summary: HRG tab, Total row, England-wide (206 providers).'),
    ('ncc_elective_inpatient_activity', 1_255_967, '#,##0', True,
     'SOURCED — same NCC 2024/25 table, Elective Inpatients activity, Total row. National '
     'completion count, used only as the blend weight below.'),
    ('ncc_daycase_cost', a['ncc_daycase_cost'], CUR, True,
     'SOURCED — same NCC 2024/25 table, Daycase column, Total row.'),
    ('ncc_daycase_activity', 7_680_341, '#,##0', True,
     'SOURCED — same NCC 2024/25 table, Daycase activity, Total row. National completion '
     'count, used only as the blend weight below.'),
    ('admitted_blended_cost', None, CUR, False,
     'Formula: (elective inpatient cost x its national activity + daycase cost x its national '
     'activity) / combined activity. RTT itself does not split "admitted" completions into '
     'daycase vs. inpatient, so this ICB\'s actual daycase/inpatient mix is unknown — bridged '
     'using the NCC\'s own NATIONAL activity mix (daycase is ~86% of national admitted '
     'activity) as the best available proxy. Flagged bridging assumption, not a guess at the '
     'unit costs themselves, which are real.'),
    ('ncc_outpatient_procedure_cost', a['ncc_outpatient_procedure_cost'], CUR, True,
     'SOURCED — same NCC 2024/25 table, Outpatient Procedures column, Total row (19,224,707 '
     'completions nationally). Used as the proxy for a non-admitted RTT pathway completion.'),
    ('icb_admitted_share', a['icb_admitted_share'], PCT, True,
     "SOURCED from this warehouse, not NCC — the 12 core C&M trusts' own trailing 6-month "
     "(Oct-2025 to Mar-2026) real RTT completed_admitted vs. completed_nonadmitted split "
     "(fact_rtt_provider_specialty_month). Genuinely this ICB's own mix, not a national figure."),
    ('unit_tariff', None, CUR, False,
     'Formula: icb_admitted_share x admitted_blended_cost + (1-icb_admitted_share) x '
     'outpatient_procedure_cost. Real inputs throughout except the one flagged national-mix '
     'bridge above — this REPLACES an earlier flat £1,500 analyst guess with a properly '
     "weighted, mostly-sourced figure once Vedant supplied NCC screenshots, 2026-08-08."),
    ('c1_inhouse', None, CUR, False,
     'Formula: unit_tariff x 100%. CORRECTED 2026-08-08 — previously 75% based on a 2022 HFMA '
     'article describing the 2022/23 ERF scheme, flagged as unverified. Verified against the '
     'actual 2025/26 NHS Payment Scheme (NHS England, "NHS provider payment mechanisms"): the '
     'marginal-rate mechanism was REMOVED — all elective activity now paid at 100% of NHSPS '
     'unit price, no floors/ceilings/marginal rates.'),
    ('c2_outsource', None, CUR, False,
     'Formula: unit_tariff x 100%. Same 2025/26 NHSPS rate applies equally to independent-'
     'sector activity — no policy-driven cost differential from in-house any more, unlike the '
     'now-corrected earlier version of this model.'),
    ('ncc_diagnostic_imaging_cost', a['ncc_diagnostic_imaging_cost'], CUR, True,
     'SOURCED — NCC 2024/25, Diagnostic Imaging column, Total row (8,573,583 tests nationally).'),
    ('ncc_audiology_cost', a['ncc_audiology_cost'], CUR, True,
     'SOURCED — NCC 2024/25, Directly Accessed Audiology, Total row (260,367 tests nationally).'),
    ('ncc_diagnostic_services_cost', a['ncc_diagnostic_services_cost'], CUR, True,
     'SOURCED — NCC 2024/25, Directly Accessed Diagnostic Services, Total row (6,180,962 tests '
     'nationally). Used as the proxy for non-audiology physiological measurement tests '
     '(echocardiography, sleep studies, neurophysiology, urodynamics) — DM01 has no more '
     'specific NCC match for these.'),
    ('dm01_imaging_share', a['dm01_imaging_share'], PCT, True,
     "SOURCED from this warehouse — the 12 core trusts' real DM01 test volume that falls in "
     "the 'imaging' category (CT/MRI/ultrasound/DEXA/barium enema), whole warehouse window."),
    ('dm01_audiology_share', a['dm01_audiology_share'], PCT, True,
     'SOURCED — DM01 audiology-assessment share of total test volume, same basis.'),
    ('dm01_other_physio_share', a['dm01_other_physio_share'], PCT, True,
     'SOURCED — DM01 share for other physiological measurement tests (cardiology, '
     'neurophysiology, sleep studies, urodynamics), same basis.'),
    ('dm01_endoscopy_share', a['dm01_endoscopy_share'], PCT, True,
     'SOURCED — DM01 endoscopy share (colonoscopy, gastroscopy, cystoscopy, flexi '
     'sigmoidoscopy), same basis. No dedicated NCC "diagnostic test" currency exists for '
     'endoscopy — most NHS endoscopies are costed as day-case/outpatient procedures, so '
     'Outpatient Procedures is used as the closest real match rather than reusing imaging.'),
    ('diag_weighted_cost', None, CUR, False,
     'Formula: DM01 category shares x their matched NCC currency cost, summed. Replaces an '
     'imaging-only proxy with one reflecting what DM01 actually measures — moves the number '
     'only slightly (£140 -> £141) because imaging is ~79% of DM01 volume, which is itself a '
     'useful confirmation that the earlier imaging-only proxy was reasonable, not a wasted step.'),
    ('diag_tests_per_pathway', a['diag_tests_per_pathway'], '0.00', True,
     'ANALYST ESTIMATE — the one fully unsourced multiplier left in this model. Plausible '
     'clinical assumption (most RTT pathways need at least one diagnostic test before a '
     'treatment decision) but not itself a published NHS ratio; no cost dataset would contain '
     'this, sourced or not.'),
    ('c3_diagnostic', None, CUR, False,
     'Formula: diag_weighted_cost x diag_tests_per_pathway. Base cost is now real and '
     'DM01-weighted; the tests-per-pathway conversion remains the weakest-sourced figure here.'),
    ('inhouse_growth_rate', a['inhouse_growth_rate'], PCT, True,
     "Max in-house capacity growth vs. current trailing 6-month completions. Loosely anchored "
     "on a national elective-activity growth ambition cited in 2022/23-era ERF reporting "
     "(HFMA) — a growth TARGET, not a payment-mechanism detail, so unaffected by the "
     "2025/26 marginal-rate correction elsewhere on this tab, but also not independently "
     "re-verified as current — treat as an analyst planning assumption, +buffer."),
    ('diag_growth_rate', a['diag_growth_rate'], PCT, True,
     'Max diagnostic-lever growth vs. current completions — more conservative than in-house, '
     'reflecting this is an indirect/bottleneck-relief lever.'),
    ('is_growth_rate', a['is_growth_rate'], PCT, True,
     "Assumed independent-sector short-term growth headroom vs. this ICB's current real "
     "outsourced RTT volume (sourced from the warehouse itself — see is_trailing_monthly "
     "below)."),
    ('bed_headroom_floor', a['bed_headroom_floor'], PCT, True,
     'KH03 G&A bed headroom (%) at/above which a provider faces no extra scaling on its '
     'in-house capacity cap; below this, the cap scales down proportionally.'),
    ('budget_total', a['default_budget'], CUR, True,
     'SCENARIO INPUT, not a sourced real ICB elective-recovery budget figure — none was '
     'publicly available. Change this cell and re-run Solver to see a different allocation.'),
    ('equity_tolerance_pp', a['equity_tolerance_pp'], PCT, True,
     "Bounded equity tolerance per Vedant's design steer (2026-08-07): higher-deprivation "
     "providers' share of total reduction can fall below their baseline breach share by at "
     "most this many percentage points — not a hard no-worsening floor."),
    ('higher_baseline_share', None, PCT, False,
     "Linked from 'Provider Inputs' — higher-deprivation providers' share of total baseline "
     "over-52wk breaches."),
    ('equity_floor_share', None, PCT, False,
     'Formula: higher_baseline_share - equity_tolerance_pp — the actual floor used in the '
     'Solver equity constraint.'),
    ('n_months', n_months, '0', True, 'Forecast horizon length, months.'),
    ('is_trailing_monthly', bundle['meta']['is_pool_monthly_total'] / a['is_growth_rate'], '#,##0', True,
     "Real, sourced from the warehouse: this ICB's actual trailing 6-month average monthly "
     "independent-sector RTT completed pathways (Mar-2026 warehouse data), NOT an estimate."),
]

AROW = {}
row = 7
for name, val, fmt, is_input, note in spec:
    AROW[name] = row
    ws.cell(row, 2, name).font = F_LABEL
    c = ws.cell(row, 3)
    if val is not None:
        c.value = val
        c.font = F_INPUT if is_input else F_FORMULA
        if is_input:
            c.fill = FILL_ASSUMPTION
    c.number_format = fmt
    n = ws.cell(row, 4, note); n.font = F_NOTE; n.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = max(15, 13 * (len(note) // 85 + 1))
    for col in (2, 3, 4):
        ws.cell(row, col).border = BORDER
    row += 1

def A(name):
    """Return an absolute reference into the Assumptions sheet, e.g. Assumptions!$C$7."""
    return f"Assumptions!$C${AROW[name]}"

def Alocal(name):
    """Same, but without the sheet prefix — for formulas written ON the Assumptions sheet itself."""
    return f"$C${AROW[name]}"

ws.cell(AROW['admitted_blended_cost'], 3,
        f"=({Alocal('ncc_elective_inpatient_cost')}*{Alocal('ncc_elective_inpatient_activity')}"
        f"+{Alocal('ncc_daycase_cost')}*{Alocal('ncc_daycase_activity')})"
        f"/({Alocal('ncc_elective_inpatient_activity')}+{Alocal('ncc_daycase_activity')})").font = F_FORMULA
ws.cell(AROW['unit_tariff'], 3,
        f"={Alocal('icb_admitted_share')}*{Alocal('admitted_blended_cost')}"
        f"+(1-{Alocal('icb_admitted_share')})*{Alocal('ncc_outpatient_procedure_cost')}").font = F_FORMULA
ws.cell(AROW['c1_inhouse'], 3, f"={Alocal('unit_tariff')}*1").font = F_FORMULA
ws.cell(AROW['c2_outsource'], 3, f"={Alocal('unit_tariff')}*1").font = F_FORMULA
ws.cell(AROW['diag_weighted_cost'], 3,
        f"={Alocal('dm01_imaging_share')}*{Alocal('ncc_diagnostic_imaging_cost')}"
        f"+{Alocal('dm01_audiology_share')}*{Alocal('ncc_audiology_cost')}"
        f"+{Alocal('dm01_other_physio_share')}*{Alocal('ncc_diagnostic_services_cost')}"
        f"+{Alocal('dm01_endoscopy_share')}*{Alocal('ncc_outpatient_procedure_cost')}").font = F_FORMULA
ws.cell(AROW['c3_diagnostic'], 3,
        f"={Alocal('diag_weighted_cost')}*{Alocal('diag_tests_per_pathway')}").font = F_FORMULA
# higher_baseline_share and equity_floor_share are wired after 'Provider Inputs' is built (below)
ws.freeze_panes = 'B7'

# ===========================================================================
# SHEET 3: Provider Inputs
# ===========================================================================
ws = wb.create_sheet('Provider Inputs')
ws.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJKL', [3, 8, 34, 12, 16, 14, 16, 15, 14, 14, 14, 14]):
    ws.column_dimensions[col].width = w

ws.cell(2, 2, '12 Core Providers — Baseline Data & Lever Caps').font = F_TITLE
ws.cell(3, 2, 'Green = sourced from nhs_warehouse.db · Black = formula on this sheet').font = F_NOTE

hdr = ['Code', 'Provider', 'LA (host)', 'LA IMD rank', 'Higher-deprivation?',
       'Baseline over-52wk breaches (18mo)', 'Trailing monthly completions',
       'Bed headroom %', 'x1 cap (in-house, /mo)', 'x2 cap (outsource, /mo)', 'x3 cap (diagnostic, /mo)']
for i, h in enumerate(hdr):
    c = ws.cell(5, 2 + i, h); c.font = F_HEADER; c.fill = FILL_HEADER
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
ws.row_dimensions[5].height = 42

start_row = 6
for i, p in enumerate(providers):
    rr = start_row + i
    ws.cell(rr, 2, p['provider_org_code']).font = F_LINK
    ws.cell(rr, 3, p['provider_org_name']).font = F_LINK
    ws.cell(rr, 4, p.get('la_name', '')).font = F_LINK
    ws.cell(rr, 5, p['la_imd_rank']).font = F_LINK
    ws.cell(rr, 6, f'=E{rr}<=$E${start_row + len(providers)}').font = F_FORMULA  # <= median row (below)
    ws.cell(rr, 7, p['baseline_over52_total']).font = F_LINK
    ws.cell(rr, 7).number_format = '#,##0'
    ws.cell(rr, 8, p['trailing_completions']).font = F_LINK
    ws.cell(rr, 8).number_format = '#,##0'
    ws.cell(rr, 9, p['bed_headroom_pct']).font = F_LINK
    ws.cell(rr, 9).number_format = PCT
    ws.cell(rr, 10, f"=H{rr}*{A('inhouse_growth_rate')}*MIN(1,I{rr}/{A('bed_headroom_floor')})")
    ws.cell(rr, 10).font = F_FORMULA; ws.cell(rr, 10).number_format = '#,##0.0'
    ws.cell(rr, 11,
            f"=G{rr}/SUM($G${start_row}:$G${start_row+len(providers)-1})"
            f"*{A('is_trailing_monthly')}*{A('is_growth_rate')}")
    ws.cell(rr, 11).font = F_FORMULA; ws.cell(rr, 11).number_format = '#,##0.0'
    ws.cell(rr, 12, f"=H{rr}*{A('diag_growth_rate')}")
    ws.cell(rr, 12).font = F_FORMULA; ws.cell(rr, 12).number_format = '#,##0.0'
    for col in range(2, 13):
        ws.cell(rr, col).border = BORDER

median_row = start_row + len(providers)
ws.cell(median_row, 3, 'MEDIAN LA IMD rank ->').font = F_LABEL
ws.cell(median_row, 5, f'=MEDIAN(E{start_row}:E{start_row+len(providers)-1})').font = F_FORMULA
ws.cell(median_row, 5).number_format = '0.0'

total_row = median_row + 2
ws.cell(total_row, 3, 'TOTALS').font = F_LABEL
for c in range(2, 13):
    ws.cell(total_row, c).fill = FILL_TOTAL
ws.cell(total_row, 7, f'=SUM(G{start_row}:G{start_row+len(providers)-1})').font = F_FORMULA
ws.cell(total_row, 7).number_format = '#,##0'

higher_share_row = total_row + 1
ws.cell(higher_share_row, 3, "Higher-deprivation providers' share of baseline breaches").font = F_LABEL
ws.cell(higher_share_row, 4,
        f'=SUMIF(F{start_row}:F{start_row+len(providers)-1},TRUE,G{start_row}:G{start_row+len(providers)-1})'
        f'/G{total_row}')
ws.cell(higher_share_row, 4).font = F_FORMULA
ws.cell(higher_share_row, 4).number_format = PCT
ws.freeze_panes = 'C6'

PI_START = start_row
PI_HIGHER_SHARE_CELL = f"'Provider Inputs'!D{higher_share_row}"

# Now wire the two Assumptions formulas that depend on Provider Inputs
wsA = wb['Assumptions']
wsA.cell(AROW['higher_baseline_share'], 3, f"={PI_HIGHER_SHARE_CELL}").font = F_LINK
wsA.cell(AROW['equity_tolerance_pp'], 3).number_format = PCT
wsA.cell(AROW['equity_floor_share'], 3,
         f"={A('higher_baseline_share').replace('Assumptions!','')}-{A('equity_tolerance_pp').replace('Assumptions!','')}").font = F_FORMULA

# ===========================================================================
# SHEET 4: Solver Model
# ===========================================================================
ws = wb.create_sheet('Solver Model')
ws.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGHIJKLMN', [3, 8, 30, 15, 20, 13, 13, 13, 13, 13, 13, 15, 15, 15]):
    ws.column_dimensions[col].width = w

ws.cell(2, 2, 'Solver Model — Decision Variables & Constraints').font = F_TITLE
ws.cell(3, 2, 'Blue cells (I:K) are the Solver decision variables — edit directly or let Excel Solver optimize them.').font = F_NOTE

hdr = ['Code', 'Provider', 'Higher-dep?', 'Baseline 18mo', 'x1 cap/mo', 'x2 cap/mo', 'x3 cap/mo',
       'x1 In-house/mo', 'x2 Outsource/mo', 'x3 Diagnostic/mo', 'Reduction 18mo', 'Residual 18mo', 'Cost 18mo (£)']
for i, h in enumerate(hdr):
    c = ws.cell(5, 2 + i, h); c.font = F_HEADER; c.fill = FILL_HEADER
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
ws.row_dimensions[5].height = 30

sm_start = 6
for i, p in enumerate(providers):
    rr = sm_start + i
    pi_row = PI_START + i
    code = p['provider_org_code']
    ws.cell(rr, 2, f"='Provider Inputs'!B{pi_row}").font = F_LINK
    ws.cell(rr, 3, f"='Provider Inputs'!C{pi_row}").font = F_LINK
    ws.cell(rr, 4, f"='Provider Inputs'!F{pi_row}").font = F_LINK
    ws.cell(rr, 5, f"='Provider Inputs'!G{pi_row}").font = F_LINK; ws.cell(rr, 5).number_format = '#,##0'
    ws.cell(rr, 6, f"='Provider Inputs'!J{pi_row}").font = F_LINK; ws.cell(rr, 6).number_format = '#,##0.0'
    ws.cell(rr, 7, f"='Provider Inputs'!K{pi_row}").font = F_LINK; ws.cell(rr, 7).number_format = '#,##0.0'
    ws.cell(rr, 8, f"='Provider Inputs'!L{pi_row}").font = F_LINK; ws.cell(rr, 8).number_format = '#,##0.0'
    rsol = ref[code]
    ws.cell(rr, 9, round(rsol['x1_inhouse_monthly'], 3)).font = F_INPUT
    ws.cell(rr, 9).fill = FILL_ASSUMPTION; ws.cell(rr, 9).number_format = '#,##0.0'
    ws.cell(rr, 10, round(rsol['x2_outsource_monthly'], 3)).font = F_INPUT
    ws.cell(rr, 10).fill = FILL_ASSUMPTION; ws.cell(rr, 10).number_format = '#,##0.0'
    ws.cell(rr, 11, round(rsol['x3_diagnostic_monthly'], 3)).font = F_INPUT
    ws.cell(rr, 11).fill = FILL_ASSUMPTION; ws.cell(rr, 11).number_format = '#,##0.0'
    ws.cell(rr, 12, f"=MIN(E{rr},{A('n_months')}*(I{rr}+J{rr}+K{rr}))")
    ws.cell(rr, 12).font = F_FORMULA; ws.cell(rr, 12).number_format = '#,##0'
    ws.cell(rr, 13, f"=E{rr}-L{rr}"); ws.cell(rr, 13).font = F_FORMULA; ws.cell(rr, 13).number_format = '#,##0'
    ws.cell(rr, 14, f"={A('n_months')}*({A('c1_inhouse')}*I{rr}+{A('c2_outsource')}*J{rr}+{A('c3_diagnostic')}*K{rr})")
    ws.cell(rr, 14).font = F_FORMULA; ws.cell(rr, 14).number_format = CUR
    for col in range(2, 15):
        ws.cell(rr, col).border = BORDER

sm_end = sm_start + len(providers) - 1
tot_row = sm_end + 2
ws.cell(tot_row, 3, 'TOTAL / OBJECTIVE').font = F_LABEL
for col in range(2, 15):
    ws.cell(tot_row, col).fill = FILL_TOTAL
ws.cell(tot_row, 5, f'=SUM(E{sm_start}:E{sm_end})').font = F_FORMULA; ws.cell(tot_row, 5).number_format = '#,##0'
ws.cell(tot_row, 12, f'=SUM(L{sm_start}:L{sm_end})').font = F_FORMULA; ws.cell(tot_row, 12).number_format = '#,##0'
ws.cell(tot_row, 13, f'=SUM(M{sm_start}:M{sm_end})').font = F_FORMULA; ws.cell(tot_row, 13).number_format = '#,##0'
ws.cell(tot_row, 14, f'=SUM(N{sm_start}:N{sm_end})').font = F_FORMULA; ws.cell(tot_row, 14).number_format = CUR

OBJ_CELL = f'L{tot_row}'
COST_CELL = f'N{tot_row}'

r2 = tot_row + 2
ws.cell(r2, 3, 'Objective cell (Solver: Set Objective, To: Max)').font = F_LABEL
ws.cell(r2, 5, f'={OBJ_CELL}').font = F_FORMULA; ws.cell(r2, 5).number_format = '#,##0'; ws.cell(r2, 5).fill = FILL_WARN
r3 = r2 + 1
ws.cell(r3, 3, '% of baseline breaches eliminated').font = F_LABEL
ws.cell(r3, 5, f'={OBJ_CELL}/E{tot_row}').font = F_FORMULA; ws.cell(r3, 5).number_format = PCT
r4 = r3 + 1
ws.cell(r4, 3, 'Total cost (18mo, £)').font = F_LABEL
ws.cell(r4, 5, f'={COST_CELL}').font = F_FORMULA; ws.cell(r4, 5).number_format = CUR
r5 = r4 + 1
ws.cell(r5, 3, 'Budget (£) — constraint: cost <= budget').font = F_LABEL
ws.cell(r5, 5, f'={A("budget_total")}').font = F_LINK; ws.cell(r5, 5).number_format = CUR
BUDGET_CELL = f'E{r5}'
r6 = r5 + 1
ws.cell(r6, 3, 'Budget slack (must be >= 0)').font = F_LABEL
ws.cell(r6, 5, f'={BUDGET_CELL}-{COST_CELL}').font = F_FORMULA; ws.cell(r6, 5).number_format = CUR
r7 = r6 + 1
ws.cell(r7, 3, "Higher-deprivation providers' reduction").font = F_LABEL
ws.cell(r7, 5, f'=SUMIF(D{sm_start}:D{sm_end},TRUE,L{sm_start}:L{sm_end})').font = F_FORMULA
ws.cell(r7, 5).number_format = '#,##0'
HIGHER_RED_CELL = f'E{r7}'
r8 = r7 + 1
ws.cell(r8, 3, 'Equity floor share (of total reduction)').font = F_LABEL
ws.cell(r8, 5, f'={A("equity_floor_share")}').font = F_LINK; ws.cell(r8, 5).number_format = PCT
EQ_FLOOR_CELL = f'E{r8}'
r9 = r8 + 1
ws.cell(r9, 3, 'Equity check cell (must be >= 0)').font = F_LABEL
ws.cell(r9, 5, f'={HIGHER_RED_CELL}-{EQ_FLOOR_CELL}*{OBJ_CELL}').font = F_FORMULA
ws.cell(r9, 5).number_format = '#,##0'
EQ_CHECK_CELL = f'E{r9}'
for rr in (r2, r3, r4, r5, r6, r7, r8, r9):
    ws.cell(rr, 5).border = BORDER

instr_row = r9 + 3
ws.cell(instr_row, 2, 'Excel Solver setup (Data > Solver — enable via File > Options > Add-ins if not visible)').font = F_LABEL
instructions = [
    f"Set Objective: ${OBJ_CELL}  ->  To: Max",
    f"By Changing Variable Cells: $I${sm_start}:$K${sm_end}",
    f"Constraint 1 (budget): ${COST_CELL} <= ${BUDGET_CELL}",
    f"Constraint 2 (equity): ${EQ_CHECK_CELL} >= 0",
    f"Constraint 3 (in-house cap): $I${sm_start}:$I${sm_end} <= $E${sm_start}:$E${sm_end}",
    f"Constraint 4 (outsource cap): $J${sm_start}:$J${sm_end} <= $F${sm_start}:$F${sm_end}",
    f"Constraint 5 (diagnostic cap): $K${sm_start}:$K${sm_end} <= $G${sm_start}:$G${sm_end}",
    f"Constraint 6 (non-negativity): $I${sm_start}:$K${sm_end} >= 0",
    "Solving Method: Simplex LP (this is a linear program — GRG Nonlinear will also work but is slower and unnecessary)",
    "Click Solve. Excel's Solver should converge to the same objective value already shown "
    "in the pre-filled cells above (a Python/scipy reference solve of the identical LP) — if "
    "it doesn't, something in the constraint setup above was mistyped, not a modeling error.",
]
rr = instr_row + 1
for line in instructions:
    ws.cell(rr, 2, f'• {line}').font = F_NORMAL
    ws.cell(rr, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=13)
    ws.row_dimensions[rr].height = 14 * (len(line) // 100 + 1)
    rr += 1

ws.freeze_panes = 'C6'

print("Objective cell:", OBJ_CELL, "| Cost cell:", COST_CELL, "| Budget cell:", BUDGET_CELL)
print("Higher-red cell:", HIGHER_RED_CELL, "| Eq floor cell:", EQ_FLOOR_CELL, "| Eq check cell:", EQ_CHECK_CELL)
print("AROW map:", AROW)

# ===========================================================================
# SHEET 5: Budget Sensitivity
# ===========================================================================
ws = wb.create_sheet('Budget Sensitivity')
ws.sheet_view.showGridLines = False
for col, w in zip('ABCDEFGH', [3, 16, 16, 16, 16, 16, 14, 14]):
    ws.column_dimensions[col].width = w
ws.cell(2, 2, 'Budget Sensitivity — 9-point sweep (reference Python solves, not live formulas)').font = F_TITLE
ws.cell(3, 2, 'Shows which lever activates at which budget level. Re-solved independently per budget with scipy — read-only reference, not Solver-linked.').font = F_NOTE

hdr = ['Budget (£, 18mo)', 'In-house units (18mo)', 'Outsource units (18mo)', 'Diagnostic units (18mo)',
       'Total reduction (breaches)', '% baseline eliminated']
for i, h in enumerate(hdr):
    c = ws.cell(5, 2 + i, h); c.font = F_HEADER; c.fill = FILL_HEADER
    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
ws.row_dimensions[5].height = 30

bs_start = 6
for i, rowd in enumerate(sens):
    rr = bs_start + i
    ws.cell(rr, 2, rowd['budget']).number_format = CUR
    ws.cell(rr, 3, rowd['total_inhouse_units']).number_format = '#,##0'
    ws.cell(rr, 4, rowd['total_outsource_units']).number_format = '#,##0'
    ws.cell(rr, 5, rowd['total_diagnostic_units']).number_format = '#,##0'
    ws.cell(rr, 6, rowd['total_reduction']).number_format = '#,##0'
    ws.cell(rr, 7, rowd['pct_reduction']).number_format = PCT
    for col in range(2, 8):
        ws.cell(rr, col).font = F_NORMAL
        ws.cell(rr, col).border = BORDER
bs_end = bs_start + len(sens) - 1

chart = LineChart()
chart.title = "Lever activation by budget level"
chart.y_axis.title = 'Units (18-month total)'
chart.x_axis.title = 'Budget (£)'
chart.width, chart.height = 24, 12
data = Reference(ws, min_col=3, max_col=5, min_row=5, max_row=bs_end)
cats = Reference(ws, min_col=2, min_row=bs_start, max_row=bs_end)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, f'B{bs_end + 3}')

note_row = bs_end + 27
ws.cell(note_row, 2, "Reading this (real NCC costs + the 2025/26 payment-scheme correction that "
        "made in-house and outsourcing cost the SAME): diagnostics (£247/unit) fills first as "
        "budget rises since it's structurally cheaper, then in-house and outsourcing (both "
        "£472/unit — genuinely tied on cost now, not one preferred over the other) activate "
        "together once diagnostic headroom exhausts around £90m ICB-wide, simply because "
        "they're the two remaining levers with capacity left. Beyond that point, total "
        "reduction plateaus around 98% regardless of further budget: every lever's CAPACITY "
        "ceiling (workforce/bed headroom, outsourcing pool) is now the binding constraint, not "
        "money. Below £90m this is a budget problem; above it, a capacity problem.").font = F_NOTE
ws.cell(note_row, 2).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=7)
ws.row_dimensions[note_row].height = 45

wb.save('/tmp/nhs_capacity_optimizer.xlsx')
print("\nSaved /tmp/nhs_capacity_optimizer.xlsx")
