"""
NHS Cheshire and Merseyside ICB: Elective Capacity Scenario and Strategy Model
Phase 5, part 3: build the deliverable workbook.

Reads the engine output (/tmp/scenario_wayfinding.json) and the early-warning
outputs, and assembles a working Excel model.

Design intent: this is a model to be used, not a report to be read. LP outputs
that Excel cannot reproduce are written as inputs and marked in blue; every
quantity that can be derived inside the workbook is written as a live formula
so the ICB can change the objective weights on the Decision Analysis tab and
watch the strategy ordering move. Anything resting on judgement rather than
computation is labelled as such on the face of the sheet.
"""
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

SRC = '/tmp/scenario_wayfinding.json'
OUT_DIR = '/tmp/scenario_out'
OUT = '/tmp/NHS_CM_Scenario_Strategy_Model.xlsx'

FONT = 'Arial'
NAVY = '1F3864'
INPUT_BLUE = Font(name=FONT, size=10, color='0000FF')
FORMULA = Font(name=FONT, size=10, color='000000')
JUDGE = Font(name=FONT, size=10, color='7F6000', italic=True)
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
H1 = Font(name=FONT, size=14, bold=True, color='FFFFFF')
H2 = Font(name=FONT, size=11, bold=True, color=NAVY)
HDR_FILL = PatternFill('solid', fgColor=NAVY)
SUB_FILL = PatternFill('solid', fgColor='D9E2F3')
YELLOW = PatternFill('solid', fgColor='FFFF00')
GREY = PatternFill('solid', fgColor='F2F2F2')
THIN = Side(style='thin', color='BFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def title(ws, text, sub, width=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(row=1, column=1, value=text)
    c.font = H1; c.fill = HDR_FILL; c.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c2 = ws.cell(row=2, column=2 - 1, value=sub)
    c2.font = Font(name=FONT, size=9, italic=True, color='595959')
    c2.alignment = Alignment(vertical='center', indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 28
    ws.sheet_view.showGridLines = False


def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=FONT, size=9, bold=True, color='FFFFFF')
        c.fill = HDR_FILL
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        c.border = BOX
    ws.row_dimensions[row].height = 32
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def put(ws, r, c, v, font=BODY, fmt=None, fill=None, wrap=False, comment=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    cell.border = BOX
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(wrap_text=wrap, vertical='top')
    if comment:
        cell.comment = Comment(comment, 'Model documentation')
    return cell


def main():
    d = json.load(open(SRC))
    grid = pd.DataFrame(d['grid'])
    matrix = pd.DataFrame(d['matrix'])
    tips = pd.DataFrame(d['tipping_points'])
    stab = pd.DataFrame(d['stability_check'])
    strategies = d['strategies']
    flags = pd.read_csv(f'{OUT_DIR}/early_warning_flags.csv')
    tri = json.load(open(f'{OUT_DIR}/diagnostic_triangulation.json'))

    wb = Workbook()

    # =====================================================================
    # 1. Read me
    # =====================================================================
    ws = wb.active
    ws.title = 'Read me'
    title(ws, 'Elective Capacity: Scenario and Strategy Model',
          'NHS Cheshire and Merseyside ICB (QYG). 12 core acute and specialist trusts, '
          '18 months to September 2027', width=6)
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 95

    rows = [
        ('What this model is for',
         'Phase 3 of this programme produced a single optimal allocation of a fixed budget across '
         'three capacity levers. That answer depends on seven parameters, four of which are analyst '
         'estimates rather than sourced figures. This model asks the question the optimiser cannot: '
         'given that uncertainty, which commitment should the ICB actually make, and what would have '
         'to be observed for that commitment to be wrong.'),
        ('What it is not',
         'It is not a forecast and it does not assign probabilities to the scenarios. Scenarios here '
         'are tests of plausibility, in the intuitive-logics tradition, not predictions.'),
        ('How to use it',
         'Start at Recommendation. Decision Analysis holds the objective weights in yellow cells. '
         'Change them and the strategy ordering recalculates. Early Warning lists the indicators to '
         'refresh monthly, and Results Grid holds the underlying numbers.'),
        ('Reading the colours',
         'Blue figures are inputs computed by the linear programme outside this workbook and cannot '
         'be reproduced by Excel formulas. Black figures are calculated live inside the workbook. '
         'Amber italics mark analyst judgement rather than computation. Yellow cells are for the '
         'user to change.'),
        ('Provenance',
         'All activity, workforce, bed and deprivation data is drawn from the project warehouse '
         '(NHS England RTT, DM01, KH03, Workforce Statistics, IMD 2019). Unit costs are from the '
         'National Cost Collection 2024/25 and the 2025/26 NHS Payment Scheme. Every parameter '
         'carries an evidence grade on the Driving Forces tab.'),
        ('The one figure with no source',
         'The diagnostic lever is priced using an assumed 1.75 diagnostic tests per unlocked '
         'pathway. No NHS dataset publishes this ratio. The scenario analysis shows the whole '
         'recommendation turns on it, so it is triangulated against observed data on the '
         'Diagnostic Triangulation tab and monitored as flag F2.'),
        ('Known limitation',
         'The linear programme has multiple equally optimal solutions whenever two levers are '
         'priced identically. Provider-level allocations are therefore not unique. Totals, '
         'constraint satisfaction, the equity figures and the delivery-risk figures are reliable. '
         'Each is pinned by a hierarchical solve that maximises volume, then equity within '
         'that, then minimises reliance on the weakest-evidenced levers within that. A specific '
         'provider-by-provider split is still not unique and must not be used as a distribution '
         'plan.'),
    ]
    r = 4
    for k, v in rows:
        put(ws, r, 2, k, font=H2, wrap=True)
        put(ws, r, 3, v, font=BODY, wrap=True)
        ws.row_dimensions[r].height = max(30, 12 * (len(v) // 95 + 1))
        r += 1

    put(ws, r + 1, 2, 'Method basis', font=H2)
    put(ws, r + 1, 3,
        'Cairns, G. and Wright, G. (2018) Scenario Thinking, 2nd ed., Palgrave Macmillan. '
        'Ch.2 basic method, Ch.5 sum-of-ranks decision analysis, Ch.8 robust strategy and flags. '
        'Sminia, H. (2026) "From Scenario Thinking to Scenario Doing: Strategic Management as '
        'Wayfinding", Futures & Foresight Science 8:e70038. Sminia, H. (2022) The Strategic '
        'Manager, 3rd ed., Ch.7 for the institutional feasibility screen.', wrap=True)
    ws.row_dimensions[r + 1].height = 60

    # =====================================================================
    # 2. Driving forces
    # =====================================================================
    ws = wb.create_sheet('Driving forces')
    title(ws, 'Stages 2-4: Driving Forces and Their Plausible Range',
          'Environmental uncertainties the ICB does not control. Equity stringency is excluded '
          'deliberately. It is a policy choice the ICB does control, so it is modelled as a '
          'dimension of strategy instead.', width=9)
    header_row(ws, 4,
               ['Cluster', 'PESTEL', 'What it is', 'Low', 'Base', 'High',
                'Low outcome', 'High outcome', 'Evidence grade and basis'],
               [22, 18, 38, 12, 12, 12, 30, 30, 62])
    r = 5
    for _, x in matrix.iterrows():
        put(ws, r, 1, x['cluster'], font=BOLD, wrap=True)
        put(ws, r, 2, x['pestel'], wrap=True)
        put(ws, r, 3, x['label'], wrap=True)
        fmt = '#,##0' if x['key'] == 'budget' else '0.000'
        put(ws, r, 4, x['low_value'], font=INPUT_BLUE, fmt=fmt)
        put(ws, r, 5, x['base_value'], font=INPUT_BLUE, fmt=fmt)
        put(ws, r, 6, x['high_value'], font=INPUT_BLUE, fmt=fmt)
        put(ws, r, 7, x['low_label'], wrap=True)
        put(ws, r, 8, x['high_label'], wrap=True)
        put(ws, r, 9, f"Grade {int(x['evidence_grade'])}. {x['evidence_basis']}", wrap=True,
            comment=x['evidence_grade_meaning'])
        ws.row_dimensions[r].height = 58
        r += 1

    put(ws, r + 1, 1, 'Evidence grading rubric', font=H2)
    r += 2
    for g, txt in sorted(d['evidence_grades'].items()):
        put(ws, r, 1, f'Grade {g}', font=BOLD)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
        put(ws, r, 2, txt, wrap=True)
        ws.row_dimensions[r].height = 26
        r += 1

    # =====================================================================
    # 3. Impact / uncertainty matrix
    # =====================================================================
    ws = wb.create_sheet('Impact-uncertainty')
    title(ws, 'Stage 5: Impact / Uncertainty Matrix',
          'The impact axis is computed, not judged. It is the swing in breaches cleared when each '
          'driving force moves across its range with all others held at base. The uncertainty axis '
          'is the evidence grade. Criticality is the product of the two, normalised.', width=8)
    header_row(ws, 4,
               ['Cluster', 'Breaches cleared at low', 'At base', 'At high',
                'Swing (impact)', 'Impact, normalised', 'Uncertainty, normalised', 'Criticality'],
               [26, 18, 16, 16, 16, 16, 18, 14])
    r = 5
    first = r
    for _, x in matrix.iterrows():
        put(ws, r, 1, x['cluster'], font=BOLD, wrap=True)
        put(ws, r, 2, x['reduction_at_low'], font=INPUT_BLUE, fmt='#,##0')
        put(ws, r, 3, x['reduction_at_base'], font=INPUT_BLUE, fmt='#,##0')
        put(ws, r, 4, x['reduction_at_high'], font=INPUT_BLUE, fmt='#,##0')
        put(ws, r, 5, f'=ABS(D{r}-B{r})', font=FORMULA, fmt='#,##0')
        r += 1
    last = r - 1
    for rr in range(first, last + 1):
        put(ws, rr, 6, f'=IF(MAX($E${first}:$E${last})=0,0,E{rr}/MAX($E${first}:$E${last}))',
            font=FORMULA, fmt='0.000')
        grade = int(matrix.iloc[rr - first]['evidence_grade'])
        put(ws, rr, 7, (grade - 1) / 3.0, font=INPUT_BLUE, fmt='0.000',
            comment=f'Evidence grade {grade}, rescaled to 0-1 as (grade-1)/3.')
        put(ws, rr, 8, f'=F{rr}*G{rr}', font=FORMULA, fmt='0.000')

    r = last + 2
    put(ws, r, 1, 'Selected scenario factors', font=H2)
    put(ws, r + 1, 1, 'Factor A', font=BOLD)
    put(ws, r + 1, 2, d['factor_a'])
    put(ws, r + 2, 1, 'Factor B', font=BOLD)
    put(ws, r + 2, 2, d['factor_b'])

    r += 4
    put(ws, r, 1, 'The most important result on this tab', font=H2)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=8)
    put(ws, r + 1, 1,
        'Three of the six candidate driving forces move the outcome by nothing at all. '
        'Independent-sector scalability, NHS in-house productivity and the size of the backlog '
        'itself have zero effect on how many breaches the programme clears, at any funding level '
        'in scope. The reason is structural: no capacity ceiling binds until the envelope reaches '
        '£31.9m (independent sector), £69.7m (diagnostic) or £107.1m (in-house), all above the '
        'range under consideration. Within that range this is not a capacity-constrained problem '
        'at all. It is a money-constrained one. Three long-running debates are therefore not '
        'live decisions, and the scenario set does not spend effort on them.', wrap=True)
    ws.row_dimensions[r + 1].height = 100

    r += 4
    put(ws, r, 1, 'Anchor-stability check', font=H2)
    header_row(ws, r + 1, ['Cluster', 'Swing at base anchor', 'Swing at diagnostic-high anchor',
                           'Criticality at base', 'Criticality at alt', 'State-dependent?'],
               [26, 20, 26, 18, 18, 16])
    rr = r + 2
    for _, x in stab.iterrows():
        put(ws, rr, 1, x['cluster'], wrap=True)
        put(ws, rr, 2, x['swing_breaches_base'], font=INPUT_BLUE, fmt='#,##0')
        put(ws, rr, 3, x['swing_breaches_alt'], font=INPUT_BLUE, fmt='#,##0')
        put(ws, rr, 4, x['criticality_base'], font=INPUT_BLUE, fmt='0.000')
        put(ws, rr, 5, x['criticality_alt'], font=INPUT_BLUE, fmt='0.000')
        put(ws, rr, 6, 'Yes' if x['state_dependent'] else 'No', font=BOLD)
        rr += 1
    put(ws, rr + 1, 1,
        'A one-at-a-time tornado is only valid at the point it is evaluated. The sweep was '
        're-run from the far corner of the most critical driver to check that no force is being '
        'recorded as inert only because of where the analysis was anchored. None flipped.',
        wrap=True)
    ws.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 1, end_column=6)
    ws.row_dimensions[rr + 1].height = 46

    # =====================================================================
    # 4. Tipping points
    # =====================================================================
    ws = wb.create_sheet('Tipping points')
    title(ws, 'Where the Answer Becomes a Different Answer',
          'A sensitivity range says how far the outcome moves. A tipping point says where the '
          'optimal lever mix changes shape. Found by bisection on the spend mix.', width=5)
    header_row(ws, 4, ['Cluster', 'Base assumption', 'Tipping point', 'Finding'],
               [26, 16, 16, 90])
    r = 5
    for _, x in tips.iterrows():
        put(ws, r, 1, x['cluster'], font=BOLD, wrap=True)
        put(ws, r, 2, x['base_value'] if pd.notna(x['base_value']) else 'n/a',
            font=INPUT_BLUE, fmt='0.000')
        put(ws, r, 3, x['tipping_point'] if pd.notna(x['tipping_point']) else 'none in range',
            font=INPUT_BLUE, fmt='0.000')
        put(ws, r, 4, x['note'], wrap=True)
        ws.row_dimensions[r].height = 30
        r += 1
    put(ws, r + 1, 1, 'Why this matters', font=H2)
    ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=4)
    put(ws, r + 2, 1,
        'Exactly one parameter in the model has a tipping point inside its plausible range, and '
        'it is the one parameter with no published source. Below 3.34 diagnostic tests per '
        'unlocked pathway, diagnostics are the cheapest route to a cleared pathway and the '
        'programme should be diagnostic-led. Above it, the diagnostic lever is dominated and the '
        'programme should buy treatment capacity instead. The working assumption of 1.75 sits '
        'below the threshold, and observed data (see Diagnostic Triangulation) sits lower still. '
        'But the entire allocation rests on which side of 3.34 the true value falls.', wrap=True)
    ws.row_dimensions[r + 2].height = 90

    # =====================================================================
    # 5. Scenarios
    # =====================================================================
    ws = wb.create_sheet('Scenarios')
    title(ws, 'Stages 6-7: The Scenario Set',
          'Four scenarios framed from the extremes of Factors A and B, plus the central case '
          'carried as a reference baseline.', width=6)
    header_row(ws, 4, ['Code', 'Name', f'Factor A: {d["factor_a"]}',
                       f'Factor B: {d["factor_b"]}', 'Funding envelope',
                       'Tests per pathway'], [10, 34, 34, 34, 18, 16])
    names = {
        'A1B1': 'Lean money, efficient diagnostics',
        'A1B2': 'Lean money, test-heavy pathways',
        'A2B1': 'Funded recovery, efficient diagnostics',
        'A2B2': 'Funded recovery, test-heavy pathways',
        'BASE': 'Central case (reference)',
    }
    r = 5
    for sc in d['scenarios']:
        put(ws, r, 1, sc['code'], font=BOLD)
        put(ws, r, 2, names.get(sc['code'], sc['code']), font=BOLD, wrap=True)
        put(ws, r, 3, sc['factor_a_state'], wrap=True)
        put(ws, r, 4, sc['factor_b_state'], wrap=True)
        put(ws, r, 5, float(sc['params']['budget']), font=INPUT_BLUE, fmt='£#,##0')
        put(ws, r, 6, float(sc['params']['diag_tests_per_pathway']), font=INPUT_BLUE, fmt='0.00')
        ws.row_dimensions[r].height = 28
        r += 1
    put(ws, r + 1, 1,
        'Scenarios are not weighted by probability. In the intuitive-logics tradition they test '
        'plausibility, and a strategy that only works in the scenario judged most likely is '
        'precisely the failure mode the method exists to catch.', wrap=True)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
    ws.row_dimensions[r + 1].height = 40

    # =====================================================================
    # 6. Strategies
    # =====================================================================
    ws = wb.create_sheet('Strategies')
    title(ws, 'Stage 3: Candidate Strategies',
          'Each strategy is a commitment expressed as a constraint set over the same linear '
          'programme, so all are evaluated on identical terms. The last three columns are '
          'analyst judgement, not computation.', width=8)
    header_row(ws, 4, ['Code', 'Strategy', 'What it commits to', 'Why it is on the list',
                       'Flexible (1-5)', 'Insurable (1-5)', 'Institutional feasibility'],
               [8, 26, 46, 46, 12, 12, 52])
    r = 5
    for s in strategies:
        put(ws, r, 1, s['code'], font=BOLD)
        put(ws, r, 2, s['name'], font=BOLD, wrap=True)
        put(ws, r, 3, s['summary'], wrap=True)
        put(ws, r, 4, s['rationale'], wrap=True)
        put(ws, r, 5, s['flexible'], font=JUDGE, fmt='0', comment=s['flex_note'])
        put(ws, r, 6, s['insurable'], font=JUDGE, fmt='0', comment=s['insure_note'])
        put(ws, r, 7, s['institutional'], font=JUDGE, wrap=True)
        ws.row_dimensions[r].height = 74
        r += 1
    put(ws, r + 1, 1,
        'Flexible and insurable follow Wright and Goodwin (2009), reported in Cairns and Wright '
        '(2018) Ch.8: is the commitment scalable up or down once made, and can its downside be '
        'capped contractually. Institutional feasibility follows Sminia (2022) Ch.7. An '
        'allocation that violates the legitimacy expectations of the NHS field will not be '
        'enacted however well it scores, so it is screened separately rather than blended into '
        'the same weighted total.', wrap=True)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=7)
    ws.row_dimensions[r + 1].height = 56

    # =====================================================================
    # 7. Results grid
    # =====================================================================
    ws = wb.create_sheet('Results grid')
    title(ws, 'Strategy x Scenario Results',
          'Every cell is a separate solve of the same linear programme. Blue figures come from '
          'the solver; black figures are derived here.', width=10)
    header_row(ws, 4, ['Scenario', 'Strategy', 'Name', 'Breaches cleared',
                       'Cleared in higher-deprivation trusts', 'Equity share',
                       'Total cost', 'Breaches per £1m', 'Delivery-risk exposure',
                       'Share of baseline cleared'],
               [12, 10, 26, 16, 22, 14, 16, 16, 18, 18])
    r = 5
    grid_first = r
    for _, x in grid.iterrows():
        put(ws, r, 1, x['scenario'], font=BOLD)
        put(ws, r, 2, x['strategy'], font=BOLD)
        put(ws, r, 3, x['strategy_name'], wrap=True)
        # Written on the same rounded basis the engine ranks on. Two strategies
        # that reach the same outcome by different routes differ only by solver
        # tolerance; ranked raw, the tie-break would be decided by floating-point
        # noise and Excel would order them differently from the engine.
        put(ws, r, 4, x['breaches_cleared'], font=INPUT_BLUE, fmt='#,##0',
            comment='Whole breaches. The single canonical basis on which strategies '
                    'are ranked, in this workbook and in the engine.')
        put(ws, r, 5, x['breaches_cleared_higher_deprivation'],
            font=INPUT_BLUE, fmt='#,##0')
        put(ws, r, 6, f'=IF(D{r}=0,0,E{r}/D{r})', font=FORMULA, fmt='0.0%')
        put(ws, r, 7, x['total_cost'], font=INPUT_BLUE, fmt='£#,##0')
        put(ws, r, 8, f'=IF(G{r}=0,0,ROUND(D{r}/(G{r}/1000000),1))', font=FORMULA, fmt='#,##0.0')
        put(ws, r, 9, x['delivery_risk_exposure'], font=INPUT_BLUE, fmt='0.0%',
            comment='Share of delivered capacity resting on the two levers whose capacity '
                    'assumptions are unsourced analyst estimates (independent-sector headroom, '
                    'diagnostic conversion). Lower is safer.')
        put(ws, r, 10, x['pct_of_baseline_cleared'], font=INPUT_BLUE, fmt='0.0%')
        r += 1
    grid_last = r - 1

    # =====================================================================
    # 8. Decision analysis: live weights
    # =====================================================================
    ws = wb.create_sheet('Decision analysis')
    title(ws, 'Ch.5 Sum-of-Ranks Decision Analysis',
          'Every strategy-scenario combination is ranked against every other on each objective, '
          'and the ranks are summed per strategy. Lower total is better. Change the weights in '
          'the yellow cells and the ordering recalculates.', width=9)

    put(ws, 4, 1, 'Objective weights', font=H2)
    put(ws, 5, 1, 'Objective', font=BOLD, fill=SUB_FILL)
    put(ws, 5, 2, 'Direction', font=BOLD, fill=SUB_FILL)
    put(ws, 5, 3, 'Weight', font=BOLD, fill=SUB_FILL)
    objs = [('Breaches cleared', 'maximise'),
            ('Cleared in higher-deprivation trusts', 'maximise'),
            ('Breaches per £1m', 'maximise'),
            ('Delivery-risk exposure', 'minimise')]
    for i, (name, direction) in enumerate(objs):
        put(ws, 6 + i, 1, name)
        put(ws, 6 + i, 2, direction)
        put(ws, 6 + i, 3, 1, font=Font(name=FONT, size=10, bold=True, color='0000FF'),
            fmt='0', fill=YELLOW)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10

    hdr = 12
    header_row(ws, hdr, ['Scenario', 'Strategy', 'Rank: breaches', 'Rank: equity',
                         'Rank: value for money', 'Rank: delivery risk', 'Weighted sum'],
               [12, 10, 16, 14, 18, 16, 14])
    r = hdr + 1
    n = grid_last - grid_first + 1
    for i in range(n):
        g = grid_first + i
        put(ws, r, 1, f"='Results grid'!A{g}", font=FORMULA)
        put(ws, r, 2, f"='Results grid'!B{g}", font=FORMULA)
        put(ws, r, 3, f"=RANK('Results grid'!D{g},'Results grid'!$D${grid_first}:$D${grid_last},0)",
            font=FORMULA, fmt='0')
        put(ws, r, 4, f"=RANK('Results grid'!E{g},'Results grid'!$E${grid_first}:$E${grid_last},0)",
            font=FORMULA, fmt='0')
        put(ws, r, 5, f"=RANK('Results grid'!H{g},'Results grid'!$H${grid_first}:$H${grid_last},0)",
            font=FORMULA, fmt='0')
        put(ws, r, 6, f"=RANK('Results grid'!I{g},'Results grid'!$I${grid_first}:$I${grid_last},1)",
            font=FORMULA, fmt='0')
        put(ws, r, 7, f'=C{r}*$C$6+D{r}*$C$7+E{r}*$C$8+F{r}*$C$9', font=FORMULA, fmt='0')
        r += 1
    rank_first, rank_last = hdr + 1, r - 1

    codes = [s['code'] for s in strategies]
    sm = r + 2
    put(ws, sm - 1, 1, 'Total by strategy (lower is better)', font=H2)
    header_row(ws, sm, ['Strategy', 'Name', 'Sum of ranks', 'Ordering'], [12, 30, 16, 12])
    for i, code in enumerate(codes):
        rr = sm + 1 + i
        nm = next(s['name'] for s in strategies if s['code'] == code)
        put(ws, rr, 1, code, font=BOLD)
        put(ws, rr, 2, nm)
        put(ws, rr, 3, f'=SUMIF($B${rank_first}:$B${rank_last},A{rr},$G${rank_first}:$G${rank_last})',
            font=FORMULA, fmt='0')
        put(ws, rr, 4, f'=RANK(C{rr},$C${sm+1}:$C${sm+len(codes)},1)', font=FORMULA, fmt='0')

    wt = sm + len(codes) + 3
    put(ws, wt, 1, 'Weight sensitivity: does the ordering survive a different priority?', font=H2)
    header_row(ws, wt + 1, ['Weighting tested', 'Resulting order, best to worst'], [40, 70])
    rr = wt + 2
    put(ws, rr, 1, 'Equal weights (as set above)', font=BOLD)
    put(ws, rr, 2, ' > '.join(x['strategy'] for x in d['by_strategy']), font=INPUT_BLUE)
    rr += 1
    for k, v in d['weight_tests'].items():
        put(ws, rr, 1, k.replace('_', ' '))
        put(ws, rr, 2, ' > '.join(v), font=INPUT_BLUE)
        rr += 1
    firsts_order = sorted({v[0] for v in d['weight_tests'].values()})
    put(ws, rr + 1, 1,
        'Two strategies take first place across these tests: ' + ', '.join(firsts_order) +
        '. S1 (and its identical twins S2 and S5) leads on every weighting of the three outcome '
        'objectives; S7 leads whenever delivery risk is weighted at twice or three times the '
        'others. The ordering is therefore NOT weight-stable, and that is the finding. The '
        'choice depends on the Board\'s appetite for exposure to unsourced delivery assumptions, '
        'and cannot be settled by the arithmetic.', wrap=True)
    ws.merge_cells(start_row=rr + 1, start_column=1, end_row=rr + 1, end_column=2)
    ws.row_dimensions[rr + 1].height = 34

    # =====================================================================
    # 9. Regret
    # =====================================================================
    ws = wb.create_sheet('Regret')
    title(ws, 'Ch.8 Minimax Regret',
          'Regret is the breaches a strategy gives up, in a given scenario, against the best that '
          'could have been done in that scenario. A minimax-regret strategy is the one whose '
          'worst case across scenarios is least bad.', width=8)
    scens = sorted(grid.scenario.unique())
    header_row(ws, 4, ['Strategy', 'Name'] + [f'Regret: {s}' for s in scens] + ['Worst-case regret'],
               [12, 28] + [14] * len(scens) + [18])
    piv = grid.pivot(index='strategy', columns='scenario', values='breaches_cleared')
    r = 5
    for code in codes:
        nm = next(s['name'] for s in strategies if s['code'] == code)
        put(ws, r, 1, code, font=BOLD)
        put(ws, r, 2, nm, wrap=True)
        for j, sc in enumerate(scens):
            best = piv[sc].max()
            put(ws, r, 3 + j, float(best - piv.loc[code, sc]), font=INPUT_BLUE, fmt='#,##0')
        c1 = get_column_letter(3)
        c2 = get_column_letter(2 + len(scens))
        put(ws, r, 3 + len(scens), f'=MAX({c1}{r}:{c2}{r})', font=FORMULA, fmt='#,##0')
        r += 1
    put(ws, r + 1, 1,
        'Read alongside the sum-of-ranks table rather than instead of it. Minimax regret looks '
        'only at breaches cleared and is indifferent to equity, value for money and delivery '
        'risk; several strategies tie at effectively zero regret because they reach the same '
        'volume by different routes. It is the tie-break, not the decision.', wrap=True)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=3 + len(scens))
    ws.row_dimensions[r + 1].height = 46

    # =====================================================================
    # 10. Early warning
    # =====================================================================
    ws = wb.create_sheet('Early warning')
    title(ws, 'Ch.8 Early-Warning Flags',
          'What to measure each month to know which scenario is actually arriving. Restricted to '
          'quantities the warehouse holds; where none exists, the gap is stated rather than '
          'filled with a proxy.', width=7)
    header_row(ws, 4, ['Flag', 'Factor', 'Indicator', 'Source', 'Current value',
                       'Trigger', 'Action if triggered'],
               [8, 26, 38, 42, 34, 34, 60])
    r = 5
    for _, x in flags.iterrows():
        put(ws, r, 1, x['flag'], font=BOLD)
        put(ws, r, 2, x['factor'], wrap=True)
        put(ws, r, 3, x['indicator'], wrap=True)
        put(ws, r, 4, x['source'], wrap=True)
        put(ws, r, 5, str(x['current_value']), font=INPUT_BLUE, wrap=True)
        put(ws, r, 6, x['trigger'], wrap=True)
        put(ws, r, 7, x['action'], wrap=True)
        fill = GREY if str(x['status']).startswith('DORMANT') else None
        if fill:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill
        ws.row_dimensions[r].height = 76
        r += 1

    # =====================================================================
    # 11. Diagnostic triangulation
    # =====================================================================
    ws = wb.create_sheet('Diagnostic triangulation')
    title(ws, 'Closing the Last Unsourced Parameter',
          'The diagnostic lever is priced on an assumed 1.75 tests per unlocked pathway, with no '
          'published source. This tab tests that assumption against observed activity.', width=4)
    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 76
    rows = [
        ('Model assumption (tests per unlocked pathway)', tri['model_assumption'],
         'Analyst estimate carried since Phase 3. No NHS dataset publishes this ratio.'),
        ('Tipping point, strategy changes above this', tri['tipping_point'],
         'Computed by bisection. Above this value the diagnostic lever is dominated by treatment '
         'capacity and the recommended strategy changes.'),
        ('Observed, trailing 12 months', round(tri['observed_trailing_12m'], 3),
         'DM01 total diagnostic activity divided by completed RTT pathways, 12 core trusts.'),
        ('Observed, full window', round(tri['observed_full_window'], 3),
         f"Across {tri['n_months_observed']} months of data."),
        ('Observed minimum', round(tri['observed_min'], 3), 'Lowest single month in the window.'),
        ('Observed maximum', round(tri['observed_max'], 3),
         'Highest single month. The observed ratio has never reached even half the tipping point.'),
    ]
    r = 4
    for k, v, note in rows:
        put(ws, r, 1, k, font=BOLD, wrap=True)
        put(ws, r, 2, v, font=INPUT_BLUE, fmt='0.000')
        put(ws, r, 3, note, wrap=True)
        ws.row_dimensions[r].height = 32
        r += 1

    r += 1
    put(ws, r, 1, 'Why this does not settle the parameter', font=H2)
    r += 1
    for cav in tri['caveats']:
        put(ws, r, 1, '•', font=BODY)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        put(ws, r, 2, cav, wrap=True)
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1
    put(ws, r, 1, 'Conclusion', font=H2)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=3)
    put(ws, r + 1, 1, tri['conclusion'], wrap=True)
    ws.row_dimensions[r + 1].height = 64

    # =====================================================================
    # 12. Recommendation
    # =====================================================================
    ws = wb.create_sheet('Recommendation')
    title(ws, 'Recommendation', 'For the Elective Recovery Programme Board', width=3)
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 108
    tied = [p for p in d['degenerate_pairs']]
    tied_codes = sorted({c for pair in tied for c in pair})
    firsts = sorted({v[0] for v in d['weight_tests'].values()})
    secs = [
        ('Recommendation',
         'Commit to S2, NHS delivery only. Deliver the elective recovery programme entirely '
         'through NHS providers, weighted to diagnostic capacity while the conversion ratio '
         'holds, with no independent-sector outsourcing.'),
        ('Why this one, and what the model does and does not settle',
         'The analysis narrows the field to three commitments that are numerically identical: '
         'S1 unconstrained optimisation, S2 NHS delivery only, and S5 equity-first produce the '
         'same result on every objective in every scenario. Neither the outsourcing lock nor the '
         'equity floor binds, so the model is genuinely indifferent between them and cannot pick '
         'one. The tie is therefore broken on the ground the model does not capture: '
         'institutional legitimacy. Of the three, S2 is the most defensible to provider boards '
         'and staff side, and commits to nothing the other two would not also do.'),
        ('The one real trade-off',
         'S7, treatment capacity only, is the genuine alternative. It never touches the two '
         'levers whose capacity assumptions are unsourced, so it carries zero delivery-risk '
         'exposure in every scenario, and it takes first place whenever delivery risk is '
         'weighted at twice or three times the other objectives. It matches S2 on volume in two '
         'of the four scenarios. But where diagnostics are efficient it gives up a great deal: '
         '26,581 fewer breaches cleared in the lean-money case and 106,324 fewer in the funded '
         'case. Choosing between S2 and S7 is a judgement about risk appetite, not a '
         'calculation, and the Board should make it as one.'),
        ('What the analysis rules out',
         'Three questions that have absorbed attention are not live decisions at these funding '
         'levels. Independent-sector scalability, NHS in-house productivity and the size of the '
         'backlog itself change the outcome by nothing, because no capacity ceiling binds below '
         'a £31.9m envelope. Within the range under consideration this is a money-constrained '
         'problem, not a capacity-constrained one.'),
        ('What the decision actually turns on',
         'One number: how many diagnostic tests it takes to unlock one additional completed '
         'pathway. Below 3.34, diagnostics are the cheapest route and the programme should be '
         'diagnostic-led. Above it, the programme should buy treatment capacity instead. The '
         'working assumption is 1.75. Observed DM01 activity per completed pathway is 0.84 and '
         'has never exceeded 0.92 in 84 months, well clear of the threshold, though the observed '
         'ratio is an average and the parameter is a marginal one, so it bounds the argument '
         'rather than settling it.'),
        ('The equity finding',
         'The equity tolerance in the Phase 3 model never binds. At every funding level examined, '
         'the entire reduction can be directed to higher-deprivation providers at no cost '
         'whatsoever to total breaches cleared, because the budget clears far less than those '
         'trusts alone are carrying. Equity here is not a trade-off to be managed. It is free, '
         'and the programme should be designed to take it.'),
        ('What would change this view',
         'Flag F2: the diagnostic conversion ratio rising above 2.00 for three consecutive months '
         'warrants review, and above 3.34 reverses the lever choice. Flag F3: the share of the '
         'diagnostic waiting list waiting over six weeks averaged 10.2% over the last twelve '
         'months against 8.1% over the preceding twelve. That step up is worth watching, because '
         'the recommended strategy assumes 10% headroom for additional diagnostic activity, but '
         'it should not be overstated. Over the full window the series trends significantly '
         'downward, and over the last 24 months there is no significant trend in either '
         'direction. Flag F1: the funding envelope is the single highest-impact uncertainty in '
         'the model and has no internal data source. It must be tracked through the finance '
         'route.'),
        ('Where the model stops',
         'Every strategy examined here is a different way of allocating a budget across three '
         'levers that already exist. At the funding levels in scope the best of them clears '
         'roughly a fifth of the projected over-52-week backlog. If the objective is to clear the '
         'backlog rather than to reduce it, no allocation of this envelope across these levers '
         'reaches that, and the question stops being an optimisation problem. That is a different '
         'kind of problem from the one this model solves, and it should be named as such rather '
         'than answered with a better-tuned version of the same answer.'),
    ]
    r = 4
    for k, v in secs:
        put(ws, r, 2, k, font=H2, wrap=True)
        put(ws, r, 3, v, font=BODY, wrap=True)
        ws.row_dimensions[r].height = max(46, 13 * (len(v) // 105 + 1))
        r += 1

    wb.save(OUT)
    print('Written', OUT)


if __name__ == '__main__':
    main()
